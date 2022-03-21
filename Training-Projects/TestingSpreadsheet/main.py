import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inventory = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name) + 1
    else:
        products_per_supplier[supplier_name] = 1

    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier.get(supplier_name) + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    if inventory < 10:
        products_under_10_inventory[int(product_number)] = int(inventory)

    inventory_price.value = inventory * price
    inventory_file.save("inventory_with_total_value.xlsx")

print("How many products per Supplier:\n" + str(products_per_supplier) + "\n")
print("Total value of inventory per supplier:\n" + str(total_value_per_supplier) + "\n")
print("Products with less than 10 in the inventory:\n" + str(products_under_10_inventory) + "\n")


